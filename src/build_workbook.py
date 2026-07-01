"""Build DSB_Group8_Workbook.xlsx by consolidating every exported answer table.

Reads the CSVs each notebook writes to its ``tables/`` folder and lays them out on
clearly-named, grouped sheets with a Contents index. Formatting: bold filled header
rows, a bold section title per block, frozen top row and auto-width columns.

Run:  python src/build_workbook.py   (or import and call build()).
Re-running always reflects the latest notebook outputs (single source of truth = the CSVs).
"""
from __future__ import annotations

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import dsb_utils as dsb  # for ROOT (relative paths only)

ROOT = dsb.ROOT
OUT = os.path.join(ROOT, "DSB_Group8_Workbook.xlsx")

# ---- styling constants ----
TITLE_FONT = Font(bold=True, size=12, color="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
THIN = Side(style="thin", color="B4C6E7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _read(qfolder: str, name: str) -> pd.DataFrame | None:
    """Read a table CSV; drop a default 0..n-1 index column; keep label columns."""
    path = os.path.join(ROOT, qfolder, "tables", name)
    if not os.path.exists(path):
        print(f"  ! missing {qfolder}/tables/{name} — skipped")
        return None
    df = pd.read_csv(path)
    first = df.columns[0]
    if str(first).startswith("Unnamed"):
        col = df[first]
        if list(col) == list(range(len(df))):       # a meaningless RangeIndex -> drop
            df = df.drop(columns=[first])
        else:                                         # real row labels -> keep, blank header
            df = df.rename(columns={first: ""})
    return df


# Layout: sheet -> [(block title, qfolder, csv name), ...]
LAYOUT = {
    "Q01-05 DataPrep": [
        ("Q1–Q2 · Dataset overview (rows & date ranges)", "part1", "Q01-02_dataset_overview.csv"),
        ("Q3 · Data-quality check (missing & duplicates)", "part1", "Q03_data_quality.csv"),
        ("Q5 · SWIFT metrics available", "part1", "Q05_swift_metrics.csv"),
    ],
    "Q06-12 Descriptive": [
        ("Q6–Q8 · Price extremes (high/low & years)", "part1", "Q06-08_price_extremes.csv"),
        ("Q10 · Top-5 Brent Oil years", "part1", "Q10_top5_brent_years.csv"),
        ("Q11 · Top-5 Gold years", "part1", "Q11_top5_gold_years.csv"),
        ("Q12 · Top-5 Silver years", "part1", "Q12_top5_silver_years.csv"),
    ],
    "Q09 Annual Averages": [
        ("Q9 · Average price by year (Brent / Gold / Silver)", "part1", "Q09_annual_average.csv"),
    ],
    "Q14-26 Trends&Currency": [
        ("Q17 · Avg annual Brent, last 10 years", "part1", "Q17_brent_last10.csv"),
        ("Q18 · Latest values (Brent / Gold / Silver)", "part1", "Q18_latest_values.csv"),
        ("Q19 · Top currencies by global payment share", "part1", "Q19_top_currencies.csv"),
        ("Q22 · Most frequent in Top-5 across reports", "part1", "Q22_top5_frequency.csv"),
        ("Q23 · Offshore RMB by economy", "part1", "Q23_offshore_rmb.csv"),
        ("Q20–Q26 · Short-answer summary", "part1", "Q_summary.csv"),
    ],
    "Q27-30 Bonus": [
        ("Q27 · Volatility comparison", "part1", "Q27_volatility.csv"),
        ("Q28 · 2020 crisis year (YoY)", "part1", "Q28_crisis_2020.csv"),
        ("Bonus B1 · Annual % change", "part1", "B1_annual_pct_change.csv"),
        ("Bonus B2 · Correlation (annual averages)", "part1", "B2_correlation.csv"),
    ],
    "PartII Research": [
        ("Part II Q2 · Event timeline", "part2", "PartII_Q2_event_timeline.csv"),
        ("Part II Q8 · Crisis comparison (Oil/Gold/Silver)", "part2", "PartII_Q8_crisis_comparison.csv"),
        ("Part II Q10 · USD vs CNY across SWIFT metrics", "part2", "PartII_Q10_usd_vs_cny.csv"),
        ("Part II Q13 · Market sensitivity table", "part2", "PartII_Q13_sensitivity.csv"),
    ],
    "Annual Summary": [
        ("Annual averages + YoY % (Oil / Gold / Silver)", "part1", "annual_summary.csv"),
    ],
}

# Human description of each sheet for the Contents page.
CONTENTS_DESC = {
    "Q01-05 DataPrep": "Part I Q1–Q5 — date ranges, row counts, data quality, SWIFT metrics",
    "Q06-12 Descriptive": "Part I Q6–Q12 — price extremes and top-5 average years",
    "Q09 Annual Averages": "Part I Q9 — full annual-average table (1915/1946–2026)",
    "Q14-26 Trends&Currency": "Part I Q14–Q26 — trend/visual outputs & SWIFT currency analytics",
    "Q27-30 Bonus": "Part I Q27–Q28 & Bonus B1–B2 — volatility, crisis, % change, correlation",
    "PartII Research": "Part II Q2/Q8/Q10/Q13 — event timeline & supporting comparison tables",
    "Annual Summary": "Reference — annual averages with year-on-year % change",
}


def _auto_width(ws):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        longest = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), 60)


def _write_block(ws, df: pd.DataFrame, title: str, row: int) -> int:
    """Write a titled table starting at 1-based `row`; return the next free row."""
    tcell = ws.cell(row=row, column=1, value=title)
    tcell.font = TITLE_FONT
    row += 1
    # header
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=row, column=j, value=str(col))
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    row += 1
    # body
    for _, rec in df.iterrows():
        for j, col in enumerate(df.columns, start=1):
            val = rec[col]
            if pd.isna(val):
                val = ""
            elif hasattr(val, "item"):
                val = val.item()
            c = ws.cell(row=row, column=j, value=val)
            c.border = BORDER
        row += 1
    return row + 2  # blank gap before next block


def build():
    wb = Workbook()
    # ---- Contents sheet (first) ----
    ws_c = wb.active
    ws_c.title = "Contents"
    ws_c["A1"] = "DSB Group 8 — Consolidated Answer Workbook"
    ws_c["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws_c["A2"] = "Global Events, Commodities and Currency Markets"
    ws_c["A2"].font = Font(italic=True, size=11)
    hdr_row = 4
    for j, h in enumerate(["Sheet", "Covers", "Contents"], start=1):
        c = ws_c.cell(row=hdr_row, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center")

    # ---- data sheets ----
    r_contents = hdr_row + 1
    for sheet_name, blocks in LAYOUT.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        row = 1
        wrote_any = False
        covers = sheet_name.split(" ")[0]
        for title, qfolder, csv in blocks:
            df = _read(qfolder, csv)
            if df is None or df.empty:
                continue
            row = _write_block(ws, df, title, row)
            wrote_any = True
        if wrote_any:
            ws.freeze_panes = "A2"  # frozen top row
            _auto_width(ws)
        # contents row
        for j, val in enumerate([sheet_name, covers, CONTENTS_DESC.get(sheet_name, "")], start=1):
            cell = ws_c.cell(row=r_contents, column=j, value=val)
            cell.border = BORDER
        r_contents += 1

    ws_c.freeze_panes = "A5"
    _auto_width(ws_c)
    ws_c.column_dimensions["C"].width = 62

    wb.save(OUT)
    print(f"saved workbook -> {os.path.relpath(OUT, ROOT)}  ({len(wb.sheetnames)} sheets)")
    print("sheets:", wb.sheetnames)


if __name__ == "__main__":
    build()
